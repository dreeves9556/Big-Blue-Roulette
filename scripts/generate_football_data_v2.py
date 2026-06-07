import openpyxl
import json
import re
from collections import defaultdict

wb = openpyxl.load_workbook('/Users/danielsmac/Documents/Football Data/kentucky_football_per_game_stats.xlsx')

players_by_id = {}
season_players = defaultdict(list)
player_stats = defaultdict(lambda: defaultdict(dict))
all_seasons = []

VALID_POSITIONS = {'QB', 'RB', 'WR', 'TE'}

PASSING_KEYS = ['Rk', 'Player', 'Pos', 'G', 'Cmp', 'Att', 'Cmp%', 'Yds', 'TD', 'TD%', 'Int', 'Int%', 'Y/A', 'AY/A', 'Y/C', 'Y/G', 'Rate', 'Awards']
RUSHING_KEYS = ['Rk', 'Player', 'Pos', 'G', 'Att', 'Yds', 'Y/A', 'TD', 'Y/G', 'Rec', 'Yds', 'Y/R', 'TD', 'Y/G', 'Plays', 'Yds', 'Avg', 'TD', 'Awards']

RUSHING_KEY_MAP = {
    'Att': 'RushAtt',
    'Yds': 'RushYds',
    'Y/A': 'RushYA',
    'TD': 'RushTD',
    'Y/G': 'RushYG',
    'Rec': 'Rec',
    'Y/R': 'YPR',
    'Plays': 'Plays',
    'Avg': 'Avg',
}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    season = str(sheet_name)
    all_seasons.append(season)
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    
    passing_start = None
    rushing_start = None
    for i, row in enumerate(rows):
        if row[0] == 'Passing':
            passing_start = i
        elif row[0] == 'Rushing & Receiving':
            rushing_start = i
    
    # Parse passing
    if passing_start is not None:
        header_idx = passing_start + 1
        if header_idx < len(rows) and rows[header_idx][0] == 'Rk':
            data_start = header_idx + 1
            end = rushing_start if rushing_start else len(rows)
            for i in range(data_start, end):
                row = rows[i]
                if not row[1]:
                    break
                name = str(row[1]).strip()
                pos = str(row[2]).strip() if row[2] else ''
                if not name or not pos:
                    continue
                pid = re.sub(r'[^a-z0-9]', '_', name.lower().replace(' ', '_').replace('-', '_').replace("'", '')).strip('_')
                if not pid:
                    continue
                if pid not in players_by_id:
                    players_by_id[pid] = {'id': pid, 'fullName': name, 'seasons': [], 'primaryPosition': pos}
                if season not in players_by_id[pid]['seasons']:
                    players_by_id[pid]['seasons'].append(season)
                if pos and not players_by_id[pid].get('primaryPosition'):
                    players_by_id[pid]['primaryPosition'] = pos
                
                stats = {'section': 'passing'}
                for j, h in enumerate(PASSING_KEYS):
                    if j < len(row) and h and row[j] is not None and h not in ('Rk', 'Player', 'Pos'):
                        try:
                            stats[h] = float(row[j]) if j >= 3 else str(row[j])
                        except:
                            stats[h] = str(row[j]) if row[j] else None
                player_stats[pid][season] = stats
                if pid not in [p['id'] for p in season_players[season]]:
                    season_players[season].append(players_by_id[pid])
    
    # Parse rushing & receiving
    if rushing_start is not None:
        header_idx = rushing_start + 1
        if header_idx < len(rows) and rows[header_idx][0] == 'Rk':
            data_start = header_idx + 1
            for i in range(data_start, len(rows)):
                row = rows[i]
                if not row[1]:
                    break
                name = str(row[1]).strip()
                pos = str(row[2]).strip() if row[2] else ''
                if not name or not pos:
                    continue
                pid = re.sub(r'[^a-z0-9]', '_', name.lower().replace(' ', '_').replace('-', '_').replace("'", '')).strip('_')
                if not pid:
                    continue
                
                # Build rushing stats with disambiguated keys
                rush_stats = {'section': 'rushing'}
                seen_yds = 0
                seen_td = 0
                seen_yg = 0
                for j, h in enumerate(RUSHING_KEYS):
                    if j < len(row) and h and row[j] is not None and h not in ('Rk', 'Player', 'Pos'):
                        key = None
                        if h == 'Yds':
                            seen_yds += 1
                            if seen_yds == 1:
                                key = 'RushYds'
                            elif seen_yds == 2:
                                key = 'RecYds'
                            elif seen_yds == 3:
                                key = 'TotalYds'
                        elif h == 'TD':
                            seen_td += 1
                            if seen_td == 1:
                                key = 'RushTD'
                            elif seen_td == 2:
                                key = 'RecTD'
                            elif seen_td == 3:
                                key = 'TotalTD'
                        elif h == 'Y/G':
                            seen_yg += 1
                            if seen_yg == 1:
                                key = 'RushYG'
                            elif seen_yg == 2:
                                key = 'RecYG'
                        else:
                            key = RUSHING_KEY_MAP.get(h, h)
                        
                        if key:
                            try:
                                rush_stats[key] = float(row[j]) if j >= 3 else str(row[j])
                            except:
                                rush_stats[key] = str(row[j]) if row[j] else None
                
                if pid in players_by_id:
                    existing = player_stats[pid].get(season, {})
                    if existing.get('section') == 'passing':
                        for k, v in rush_stats.items():
                            if k != 'section':
                                existing[k] = v
                        existing['section'] = 'both'
                        player_stats[pid][season] = existing
                    else:
                        player_stats[pid][season] = rush_stats
                        if season not in players_by_id[pid]['seasons']:
                            players_by_id[pid]['seasons'].append(season)
                        if pid not in [p['id'] for p in season_players[season]]:
                            season_players[season].append(players_by_id[pid])
                else:
                    if pid not in players_by_id:
                        players_by_id[pid] = {'id': pid, 'fullName': name, 'seasons': [], 'primaryPosition': pos}
                    if season not in players_by_id[pid]['seasons']:
                        players_by_id[pid]['seasons'].append(season)
                    if pos and not players_by_id[pid].get('primaryPosition'):
                        players_by_id[pid]['primaryPosition'] = pos
                    player_stats[pid][season] = rush_stats
                    if pid not in [p['id'] for p in season_players[season]]:
                        season_players[season].append(players_by_id[pid])

# Sort seasons
all_seasons = sorted(all_seasons, key=lambda s: int(s))

# Sort each player's seasons
for p in players_by_id.values():
    p['seasons'] = sorted(p['seasons'], key=lambda s: int(s))

# Only keep players with valid positions
valid_players = {pid: p for pid, p in players_by_id.items() if p['primaryPosition'] in VALID_POSITIONS}

# Filter season_players to only valid positions
for season in season_players:
    season_players[season] = [p for p in season_players[season] if p['primaryPosition'] in VALID_POSITIONS]

# Generate footballPlayers.js
players_js_lines = [
    '// Kentucky Wildcats Football Player Dataset',
    '// Source: kentucky_football_per_game_stats.xlsx',
    '',
    'export const footballPlayers = [',
]

for pid in sorted(valid_players.keys()):
    p = valid_players[pid]
    seasons_str = ', '.join(f'"{s}"' for s in p['seasons'])
    players_js_lines.append(f'  {{')
    players_js_lines.append(f'    "id": "{pid}",')
    players_js_lines.append(f'    "fullName": "{p["fullName"]}",')
    players_js_lines.append(f'    "seasons": [{seasons_str}],')
    players_js_lines.append(f'    "primaryPosition": "{p["primaryPosition"]}"')
    players_js_lines.append(f'  }},')

if len(players_js_lines) > 4:
    players_js_lines[-1] = players_js_lines[-1].rstrip(',')

players_js_lines.append('];')
players_js_lines.append('')
players_js_lines.append("export const FOOTBALL_POSITIONS = ['QB', 'RB', 'WR1', 'WR2', 'TE'];")
players_js_lines.append('')
players_js_lines.append('export const FOOTBALL_POSITION_LABELS = {')
players_js_lines.append("  QB: 'Quarterback',")
players_js_lines.append("  RB: 'Running Back',")
players_js_lines.append("  WR1: 'Wide Receiver 1',")
players_js_lines.append("  WR2: 'Wide Receiver 2',")
players_js_lines.append("  TE: 'Tight End',")
players_js_lines.append('};')
players_js_lines.append('')
players_js_lines.append('// All unique seasons across all players, sorted')
players_js_lines.append('export const footballAllSeasons = [...new Set(footballPlayers.flatMap((p) => p.seasons))].sort((a, b) => parseInt(a) - parseInt(b));')
players_js_lines.append('')
players_js_lines.append('// Map season -> players who played in that season')
players_js_lines.append('export const footballSeasonPlayersMap = Object.fromEntries(')
players_js_lines.append('  footballAllSeasons.map((season) => [')
players_js_lines.append('    season,')
players_js_lines.append('    footballPlayers.filter((p) => p.seasons.includes(season)),')
players_js_lines.append('  ])')
players_js_lines.append(');')
players_js_lines.append('')

with open('/Users/danielsmac/Documents/Kentucky Basketball Team Draft/src/data/footballPlayers.js', 'w') as f:
    f.write('\n'.join(players_js_lines))

# Generate footballPlayerSeasonStats.js
stats_obj = {}
for pid in valid_players:
    if pid in player_stats:
        stats_obj[pid] = {}
        for season in player_stats[pid]:
            stats = player_stats[pid][season]
            clean_stats = {}
            for k, v in stats.items():
                if k in ('section',):
                    continue
                if v is not None:
                    clean_stats[k] = v
            stats_obj[pid][season] = clean_stats

stats_js_lines = [
    '// Kentucky Wildcats Football Player Season Stats',
    '// Per-game stats by player ID and season',
    '',
    'export const footballPlayerSeasonStatsById = ' + json.dumps(stats_obj, indent=2) + ';',
    '',
]

with open('/Users/danielsmac/Documents/Kentucky Basketball Team Draft/src/data/footballPlayerSeasonStats.js', 'w') as f:
    f.write('\n'.join(stats_js_lines))

print(f'Generated footballPlayers.js with {len(valid_players)} players')
print(f'Generated footballPlayerSeasonStats.js with {len(stats_obj)} players with stats')
print(f'Seasons: {len(all_seasons)} ({all_seasons[0]} - {all_seasons[-1]})')

from collections import Counter
pos_counts = Counter(p['primaryPosition'] for p in valid_players.values())
print('Valid positions:', dict(pos_counts))

# Verify sample QB
for pid, p in valid_players.items():
    if p['primaryPosition'] == 'QB' and '2024' in p['seasons']:
        print('Sample QB 2024:', p)
        print('Stats:', stats_obj.get(pid, {}).get('2024'))
        break

# Verify sample RB
for pid, p in valid_players.items():
    if p['primaryPosition'] == 'RB' and '2024' in p['seasons']:
        print('Sample RB 2024:', p)
        print('Stats:', stats_obj.get(pid, {}).get('2024'))
        break

# Verify sample WR
for pid, p in valid_players.items():
    if p['primaryPosition'] == 'WR' and '2024' in p['seasons']:
        print('Sample WR 2024:', p)
        print('Stats:', stats_obj.get(pid, {}).get('2024'))
        break

# Verify sample TE
for pid, p in valid_players.items():
    if p['primaryPosition'] == 'TE' and '2024' in p['seasons']:
        print('Sample TE 2024:', p)
        print('Stats:', stats_obj.get(pid, {}).get('2024'))
        break
