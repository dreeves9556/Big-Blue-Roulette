import openpyxl
import json
import re
from collections import defaultdict

wb = openpyxl.load_workbook('/Users/danielsmac/Documents/Football Data/kentucky_football_per_game_stats.xlsx')

players_by_id = {}
season_players = defaultdict(list)
player_stats = defaultdict(lambda: defaultdict(dict))
all_seasons = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    season = str(sheet_name)
    all_seasons.append(season)
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    
    passing_start = None
    rushing_start = None
    for i, row in enumerate(rows):
        if row[0] == 'Passing':
            passing_start = i
        elif row[0] == 'Rushing & Receiving':
            rushing_start = i
    
    # Parse passing
    if passing_start is not None:
        header_idx = passing_start + 1
        if header_idx < len(rows) and rows[header_idx][0] == 'Rk':
            headers = rows[header_idx]
            data_start = header_idx + 1
            end = rushing_start if rushing_start else len(rows)
            for i in range(data_start, end):
                row = rows[i]
                if not row[1]:
                    break
                name = str(row[1]).strip()
                pos = str(row[2]).strip() if row[2] else ''
                if not name or not pos:
                    continue
                pid = re.sub(r'[^a-z]', '_', name.lower().replace(' ', '_').replace('-', '_').replace("'", '')).strip('_')
                if not pid:
                    continue
                if pid not in players_by_id:
                    players_by_id[pid] = {'id': pid, 'fullName': name, 'seasons': [], 'primaryPosition': pos}
                if season not in players_by_id[pid]['seasons']:
                    players_by_id[pid]['seasons'].append(season)
                if pos and not players_by_id[pid].get('primaryPosition'):
                    players_by_id[pid]['primaryPosition'] = pos
                
                stats = {'section': 'passing'}
                for j, h in enumerate(headers):
                    if h and row[j] is not None:
                        try:
                            stats[h] = float(row[j]) if j >= 3 else str(row[j])
                        except:
                            stats[h] = str(row[j]) if row[j] else None
                player_stats[pid][season] = stats
                if pid not in [p['id'] for p in season_players[season]]:
                    season_players[season].append(players_by_id[pid])
    
    # Parse rushing & receiving
    if rushing_start is not None:
        header_idx = rushing_start + 1
        if header_idx < len(rows) and rows[header_idx][0] == 'Rk':
            headers = rows[header_idx]
            data_start = header_idx + 1
            for i in range(data_start, len(rows)):
                row = rows[i]
                if not row[1]:
                    break
                name = str(row[1]).strip()
                pos = str(row[2]).strip() if row[2] else ''
                if not name or not pos:
                    continue
                pid = re.sub(r'[^a-z]', '_', name.lower().replace(' ', '_').replace('-', '_').replace("'", '')).strip('_')
                if not pid:
                    continue
                
                if pid in players_by_id:
                    existing = player_stats[pid].get(season, {})
                    if existing.get('section') == 'passing':
                        for j, h in enumerate(headers):
                            if h and row[j] is not None:
                                try:
                                    existing[h] = float(row[j]) if j >= 3 else str(row[j])
                                except:
                                    existing[h] = str(row[j]) if row[j] else None
                        existing['section'] = 'both'
                        player_stats[pid][season] = existing
                    else:
                        stats = {'section': 'rushing'}
                        for j, h in enumerate(headers):
                            if h and row[j] is not None:
                                try:
                                    stats[h] = float(row[j]) if j >= 3 else str(row[j])
                                except:
                                    stats[h] = str(row[j]) if row[j] else None
                        player_stats[pid][season] = stats
                        if season not in players_by_id[pid]['seasons']:
                            players_by_id[pid]['seasons'].append(season)
                        if pid not in [p['id'] for p in season_players[season]]:
                            season_players[season].append(players_by_id[pid])
                else:
                    if pid not in players_by_id:
                        players_by_id[pid] = {'id': pid, 'fullName': name, 'seasons': [], 'primaryPosition': pos}
                    if season not in players_by_id[pid]['seasons']:
                        players_by_id[pid]['seasons'].append(season)
                    if pos and not players_by_id[pid].get('primaryPosition'):
                        players_by_id[pid]['primaryPosition'] = pos
                    
                    stats = {'section': 'rushing'}
                    for j, h in enumerate(headers):
                        if h and row[j] is not None:
                            try:
                                stats[h] = float(row[j]) if j >= 3 else str(row[j])
                            except:
                                stats[h] = str(row[j]) if row[j] else None
                    player_stats[pid][season] = stats
                    if pid not in [p['id'] for p in season_players[season]]:
                        season_players[season].append(players_by_id[pid])

print(f'Total unique players: {len(players_by_id)}')
print(f'Total seasons: {len(all_seasons)}')

sample = list(players_by_id.values())[:5]
for p in sample:
    print(p)
print('---')

from collections import Counter
pos_counts = Counter(p['primaryPosition'] for p in players_by_id.values())
print('Positions:', dict(pos_counts))

# Check a QB's stats
for pid, p in players_by_id.items():
    if p['primaryPosition'] == 'QB':
        print('Sample QB:', p)
        for season, stats in player_stats[pid].items():
            print(f'  {season}: {stats}')
        break

# Check a WR's stats
for pid, p in players_by_id.items():
    if p['primaryPosition'] == 'WR':
        print('Sample WR:', p)
        for season, stats in player_stats[pid].items():
            print(f'  {season}: {stats}')
        break
